import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import os
import tempfile
from io import BytesIO

# ===================== 新增：密码验证 =====================
def check_password():
    def password_entered():
        if st.session_state["password"] == "WXqwer1234@":
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.title("🔒 请输入访问密码")
        st.text_input(
            "密码",
            type="password",
            on_change=password_entered,
            key="password"
        )
        return False
    elif not st.session_state["password_correct"]:
        st.title("🔒 请输入访问密码")
        st.text_input(
            "密码",
            type="password",
            on_change=password_entered,
            key="password"
        )
        st.error("密码错误，请重试！")
        return False
    else:
        return True

if not check_password():
    st.stop()

# ===================== 原有代码（无需修改） =====================
# 设置页面标题和布局
st.set_page_config(page_title="客户对账单生成工具", layout="wide")
st.title("📊 客户对账单自动生成工具")
st.divider()

# 定义生成对账单的核心函数（返回文件字节流，用于下载）
def create_settlement_file(year, month, customer_name, amount):
    """生成单个对账单，返回文件字节流"""
    wb = Workbook()
    ws = wb.active
    ws.title = "结算单"
    
    # 设置列宽（所有列宽20）
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 20
    
    # 边框和对齐样式
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_center_alignment = Alignment(horizontal='right', vertical='center')
    left_center_alignment = Alignment(horizontal='left', vertical='center')
    
    # 填充模板内容
    # 第1行：结算执行单编号（右居中）
    ws.merge_cells('A1:H1')
    ws['A1'] = "结算执行单编号："
    ws['A1'].alignment = right_center_alignment
    ws['A1'].border = thin_border
    
    # 第2行：标题
    ws.merge_cells('A2:H2')
    ws['A2'] = f"{year}年{month}月无锡凌恒网络技术有限公司《结算单》"
    ws['A2'].alignment = center_alignment
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].border = thin_border
    
    # 第3行：合同信息/甲方开票信息表头
    ws.merge_cells('A3:D3')
    ws['A3'] = "合同信息"
    ws['A3'].alignment = center_alignment
    ws['A3'].font = Font(bold=True)
    ws['A3'].border = thin_border
    ws.merge_cells('E3:H3')
    ws['E3'] = "甲方开票信息"
    ws['E3'].alignment = center_alignment
    ws['E3'].font = Font(bold=True)
    ws['E3'].border = thin_border
    
    # 第4行：合同名称/公司名称
    ws['A4'] = "合同名称"
    ws['A4'].alignment = center_alignment
    ws['A4'].border = thin_border
    ws.merge_cells('B4:D4')
    ws['B4'] = "数据推广协议"
    ws['B4'].alignment = center_alignment
    ws['B4'].border = thin_border
    ws['E4'] = "公司名称"
    ws['E4'].alignment = center_alignment
    ws['E4'].border = thin_border
    ws.merge_cells('F4:H4')
    ws['F4'] = customer_name
    ws['F4'].alignment = center_alignment
    ws['F4'].border = thin_border
    
    # 第5行：甲方主体/发票类型
    ws['A5'] = "甲方主体"
    ws['A5'].alignment = center_alignment
    ws['A5'].border = thin_border
    ws.merge_cells('B5:D5')
    ws['B5'] = customer_name
    ws['B5'].alignment = center_alignment
    ws['B5'].border = thin_border
    ws['E5'] = "发票类型"
    ws['E5'].alignment = center_alignment
    ws['E5'].border = thin_border
    ws.merge_cells('F5:H5')
    ws['F5'] = "增值税专用发票"
    ws['F5'].alignment = center_alignment
    ws['F5'].border = thin_border
    
    # 第6行：乙方主体/发票内容
    ws['A6'] = "乙方主体"
    ws['A6'].alignment = center_alignment
    ws['A6'].border = thin_border
    ws.merge_cells('B6:D6')
    ws['B6'] = "无锡凌恒网络技术有限公司"
    ws['B6'].alignment = center_alignment
    ws['B6'].border = thin_border
    ws['E6'] = "发票内容"
    ws['E6'].alignment = center_alignment
    ws['E6'].border = thin_border
    ws.merge_cells('F6:H6')
    ws['F6'] = "广告发布费"
    ws['F6'].alignment = center_alignment
    ws['F6'].border = thin_border
    
    # 第7行：已结算项目表头
    ws.merge_cells('A7:H7')
    ws['A7'] = "已结算项目"
    ws['A7'].alignment = center_alignment
    ws['A7'].font = Font(bold=True)
    ws['A7'].border = thin_border
    
    # 第8行：列名
    ws['A8'] = "序号"
    ws['A8'].alignment = center_alignment
    ws['A8'].border = thin_border
    ws['B8'] = "日期"
    ws['B8'].alignment = center_alignment
    ws['B8'].border = thin_border
    ws['C8'] = "结算项目"
    ws['C8'].alignment = center_alignment
    ws['C8'].border = thin_border
    ws['D8'] = "结算方式"
    ws['D8'].alignment = center_alignment
    ws['D8'].border = thin_border
    ws['E8'] = "配送方式"
    ws['E8'].alignment = center_alignment
    ws['E8'].border = thin_border
    ws['F8'] = "结算金额"
    ws['F8'].alignment = center_alignment
    ws['F8'].border = thin_border
    ws['G8'] = "小计"
    ws['G8'].alignment = center_alignment
    ws['G8'].border = thin_border
    ws['H8'] = "备注"
    ws['H8'].alignment = center_alignment
    ws['H8'].border = thin_border
    
    # 第9行：已结算项目数据（巨量引擎）
    ws['A9'] = 1
    ws['A9'].alignment = center_alignment
    ws['A9'].border = thin_border
    ws['B9'] = f"{year}年{month}月"
    ws['B9'].alignment = center_alignment
    ws['B9'].border = thin_border
    ws['C9'] = "巨量引擎"
    ws['C9'].alignment = center_alignment
    ws['C9'].border = thin_border
    ws['D9'] = "预存"
    ws['D9'].alignment = center_alignment
    ws['D9'].border = thin_border
    ws['E9'] = "立即充值"
    ws['E9'].alignment = center_alignment
    ws['E9'].border = thin_border
    ws['F9'] = amount
    ws['F9'].alignment = center_alignment
    ws['F9'].border = thin_border
    ws['G9'] = amount
    ws['G9'].alignment = center_alignment
    ws['G9'].border = thin_border
    ws['H9'] = ""
    ws['H9'].alignment = center_alignment
    ws['H9'].border = thin_border
    
    # 第10行：合计
    ws.merge_cells('A10:F10')
    ws['A10'] = "合计"
    ws['A10'].alignment = center_alignment
    ws['A10'].border = thin_border
    ws['G10'] = amount
    ws['G10'].alignment = center_alignment
    ws['G10'].border = thin_border
    ws.merge_cells('H10:H10')
    ws['H10'].border = thin_border
    
    # 第11-12行：乙方收款信息（合并+左居中）
    ws.merge_cells('A11:H12')
    ws['A11'] = f"""乙方收款信息：
户名：无锡凌恒网络技术有限公司
开户行：中国农业银行股份有限公司无锡新城支行
账号：10655501040011319"""
    ws['A11'].alignment = left_center_alignment
    ws['A11'].border = thin_border
    
    # 第13行：备注
    ws.merge_cells('A13:H13')
    ws['A13'] = "备注：甲乙双方审核确认以上信息没有问题后，请盖章并签名，并把原件快递给对方，乙方将根据以上确认的最终信息开具发票并快递给甲方。"
    ws['A13'].alignment = center_alignment
    ws['A13'].border = thin_border
    
    # 第14行：甲乙双方（合并+左居中）
    ws.merge_cells('A14:D14')
    ws['A14'] = f"甲方：{customer_name}（盖章）"
    ws['A14'].alignment = left_center_alignment
    ws['A14'].border = thin_border
    ws.merge_cells('E14:H14')
    ws['E14'] = "乙方：无锡凌恒网络技术有限公司（盖章）"
    ws['E14'].alignment = left_center_alignment
    ws['E14'].border = thin_border
    
    # 第15行：时间
    ws.merge_cells('A15:D15')
    ws['A15'] = "时间："
    ws['A15'].alignment = center_alignment
    ws['A15'].border = thin_border
    ws.merge_cells('E15:H15')
    ws['E15'] = "时间："
    ws['E15'].alignment = center_alignment
    ws['E15'].border = thin_border
    
    # 设置行高
    for row_num in range(1, 11):
        ws.row_dimensions[row_num].height = 20
    for row_num in range(11, 16):
        ws.row_dimensions[row_num].height = 50
    
    # 保存到字节流（用于下载）
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ===================== Streamlit 页面界面 =====================
# 1. 上传文件区域
st.subheader("📤 第一步：上传回款汇总Excel文件")
uploaded_file = st.file_uploader(
    label="请上传汇总表（格式：年份、月份、客户名称、回款金额）",
    type=["xlsx"],
    help="汇总表第一列：年份，第二列：月份，第三列：客户名称，第四列：回款金额"
)

# 2. 生成按钮区域
st.subheader("⚡ 第二步：生成对账单")
generate_btn = st.button("点击生成对账单", type="primary", disabled=not uploaded_file)

# 3. 生成逻辑执行
if generate_btn and uploaded_file:
    try:
        # 读取上传的汇总文件
        wb_summary = load_workbook(uploaded_file, data_only=True)
        ws_summary = wb_summary.active
        
        st.divider()
        st.subheader("✅ 生成结果（点击下载）")
        
        # 遍历汇总数据生成对账单，并提供下载按钮
        for row in ws_summary.iter_rows(min_row=2, values_only=True):
            year, month, customer_name, amount = row
            if not all([year, month, customer_name, amount]):
                st.warning(f"跳过空行：{row}")
                continue
            
            # 生成对账单字节流
            file_data = create_settlement_file(year, month, customer_name, amount)
            # 生成下载按钮
            filename = f"【结算单】{customer_name}-{year}-{month}.xlsx"
            st.download_button(
                label=f"下载 {filename}",
                data=file_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        wb_summary.close()
        st.success("所有对账单生成完成，点击按钮即可下载到你的电脑！")
    
    except Exception as e:
        st.error(f"生成失败：{str(e)}")

# 4. 操作说明
st.divider()
st.subheader("📖 操作说明")
st.markdown("""
1. 汇总Excel格式要求（第一行为列名）：
   | 年份 | 月份 | 客户名称 | 回款金额 |
   |------|------|----------|----------|
   | 2026 | 7    | 客户A    | 50000    |
2. 点击「下载」按钮后，文件会直接保存到你的电脑「下载」文件夹；
3. 无需安装任何软件，直接在线使用！
""")
